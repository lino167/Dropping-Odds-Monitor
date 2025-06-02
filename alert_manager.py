import logging
import os
from openpyxl import load_workbook
from excel_utils import registrar_alerta_excel

# Variáveis globais para rastrear alertas
sent_alerts = set()


def check_alerts(unified_data, game_id, match_url, df_total):
    """
    Verifica condições para alertas e retorna mensagens formatadas.
    Só envia se o game_id não estiver na memória e nem registrado na planilha.
    """
    global sent_alerts
    alerts = []

    if unified_data.empty:
        logging.info(
            f"Nenhum dado unificado disponível para verificar alertas no jogo {game_id}."
        )
        return alerts

    # Verifica se o alerta já foi enviado (na memória ou na planilha)
    nome_arquivo = "alertas.xlsx"
    if os.path.exists(nome_arquivo):
        workbook = load_workbook(nome_arquivo)
        sheet = workbook.active
        game_ids_registrados = [
            str(row[0].value) for row in sheet.iter_rows(min_row=2, max_col=1)
        ]
        if str(game_id) in game_ids_registrados:
            logging.info(
                f"Alerta para game_id {game_id} já registrado na planilha. Ignorando."
            )
            return alerts

    # Pega informações da tabela unificada
    league = (
        unified_data["league"].iloc[0] if "league" in unified_data.columns else "N/A"
    )
    home_team = (
        unified_data["home_team"].iloc[0]
        if "home_team" in unified_data.columns
        else "N/A"
    )
    away_team = (
        unified_data["away_team"].iloc[0]
        if "away_team" in unified_data.columns
        else "N/A"
    )

    # Identifica o favorito (primeira linha válida)
    favorite = None
    favorite_initial = None
    favorite_nome = None
    for _, row in unified_data.iterrows():
        home_odd = float(row.get("home", 0))
        away_odd = float(row.get("away", 0))
        if home_odd > 1.01 and away_odd > 1.01:
            if home_odd < away_odd:
                favorite = "home"
                favorite_initial = home_odd
                favorite_nome = home_team
            else:
                favorite = "away"
                favorite_initial = away_odd
                favorite_nome = away_team
            break 

    if not favorite or not favorite_initial or favorite_initial <= 1.01:
        logging.info(f"Favorito não encontrado ou inválido para o jogo {game_id}.")
        return alerts  # Não achou favorito válido

    # Pega o handicap e odd inicial da primeira linha da tabela unificada
    handicap_inicial = unified_data.iloc[0].get("handicap", "N/A")
    over_inicial = unified_data.iloc[0].get("over", "N/A")

    for _, row in unified_data.iterrows():
        # 1. Checa o drop da linha de gols primeiro
        drop_gols = row.get("drop", 0)
        try:
            drop_gols_float = float(drop_gols)
        except Exception:
            continue
        if drop_gols_float < 0.60:
            continue

        # 2. Verifica cartões vermelhos e pênaltis
        red_card = row.get("red_card", "")
        penalty = row.get("penalty", "")
        if red_card != "0-0" or penalty != "0-0":
            continue

        # 3. Placar
        score = row.get("score", "")
        try:
            gols_home, gols_away = map(int, str(score).split("-"))
        except Exception:
            continue

        # 4. Só alerta se favorito não está vencendo (empatando ou perdendo)
        if (favorite == "home" and gols_home > gols_away) or (
            favorite == "away" and gols_away > gols_home
        ):
            continue

        # 5. Odds atuais válidas
        odd_atual = float(row.get(favorite, 0))
        if odd_atual <= 1.01:
            continue

        # 6. Minuto do jogo deve ser antes do 80
        time_min = row.get("time", 0)
        try:
            time_min_int = int(time_min)
        except Exception:
            continue
        if time_min_int >= 80:
            continue

        # 7. Cálculo da variação
        if odd_atual < favorite_initial:
            variacao = round(
                ((favorite_initial - odd_atual) / favorite_initial) * 100, 2
            )
            tipo = "Queda odd favorito"
            emoji = "📉"
        elif odd_atual > favorite_initial:
            variacao = round(((odd_atual - favorite_initial) / odd_atual) * 100, 2)
            tipo = "Subida odd favorito"
            emoji = "📈"
        else:
            continue  # Sem variação

        # 8. Drop do total (para mostrar no alerta se for relevante)
        info_drop_total = f"{(drop_gols_float)*100}%"

        # Monta alerta
        alerta = [
            game_id,
            league,
            f"{home_team} - {away_team}",
            favorite_nome, 
            favorite_initial,
            odd_atual,
            variacao,
            handicap_inicial,
            over_inicial,
            row.get("handicap", "N/A"),
            row.get("over", "N/A"),
            row.get("drop", "N/A"),
            score,
            time_min,
            match_url,
        ]
        if registrar_alerta_excel(game_id, alerta):
            sent_alerts.add(game_id)
            message = (
                f"🤖 Alerta automático para o jogo <b>{home_team} x {away_team} ({league})</b>:\n\n"
                f"📉 Detectado um drop significativa na odd da linha de gols: {info_drop_total}\n"
                f"📊 Linha de gols inicial: {handicap_inicial} → (Odd: {over_inicial})\n"
                f"⚡ Linha de gols atual: {row.get('handicap', 'N/A')} → (Odd: {row.get('over', 'N/A')})\n"
                f"⭐ O favorito é o <b>{favorite_nome}</b> com {'queda' if 'Queda' in tipo else 'subida'} "
                f"na odd de: {abs(variacao)}% ({favorite_initial} → {odd_atual}).\n"
                f"🔢 Placar atual: {score}\n"
                f"⏱️ Minuto: {time_min}\n\n"
                f"Para mais detalhes, acesse: <a href='{match_url}'>Aqui</a>\n\n"
                f"⚠️ Lembre-se: este é um alerta automatizado, não é recomendação de aposta."
            )
            alerts.append(message)
    return alerts
