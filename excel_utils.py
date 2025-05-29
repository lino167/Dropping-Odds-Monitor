import os
import logging
from openpyxl import Workbook, load_workbook
from config import EXCEL_COLUMNS

def criar_planilha_excel(nome_arquivo="alertas.xlsx"):
    """Cria uma planilha Excel com as colunas especificadas, caso ela não exista."""
    if not os.path.exists(nome_arquivo):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Alertas"
        sheet.append(EXCEL_COLUMNS)
        workbook.save(nome_arquivo)
        logging.info(f"Planilha '{nome_arquivo}' criada com sucesso.")
    else:
        logging.info(f"Planilha '{nome_arquivo}' já existe.")

def registrar_alerta_excel(game_id, alerta, nome_arquivo="alertas.xlsx"):
    """Registra um alerta na planilha Excel, caso o game_id ainda não esteja registrado."""
    if not os.path.exists(nome_arquivo):
        criar_planilha_excel(nome_arquivo)

    workbook = load_workbook(nome_arquivo)
    sheet = workbook.active

    # Verifica se o game_id já está registrado
    game_ids_registrados = [row[0].value for row in sheet.iter_rows(min_row=2, max_col=1)]
    if game_id in game_ids_registrados:
        logging.info(f"Alerta para game_id {game_id} já registrado. Ignorando.")
        return False

    sheet.append(alerta)
    workbook.save(nome_arquivo)
    logging.info(f"Alerta registrado para game_id {game_id} na planilha '{nome_arquivo}'.")
    return True