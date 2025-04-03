import os
import json
import pandas as pd
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink
from config.settings import SENT_GAMES_FILE

def save_to_json(data, filename='temp_data.json'):
    with open(filename, 'w') as json_file:
        json.dump(data, json_file, indent=4)

def load_from_json(filename='temp_data.json'):
    with open(filename, 'r') as json_file:
        return json.load(json_file)

def delete_json(filename='temp_data.json'):
    if os.path.exists(filename):
        os.remove(filename)

def initialize_excel_file():
    global sent_games_df
    try:
        if not os.path.exists(SENT_GAMES_FILE):
            sent_games_df = pd.DataFrame(columns=[
                'game_id', 'Liga', 'Team A', 'Team B', 'Placar', 'Tempo', 'Odd', 'Linha de gol over', 'Drop', 'Link', 'Tipo de Alerta'
            ])
            sent_games_df.to_excel(SENT_GAMES_FILE, index=False)
            logging.info(f"Planilha criada: {SENT_GAMES_FILE}")
        else:
            sent_games_df = pd.read_excel(SENT_GAMES_FILE)

            if 'game_id' not in sent_games_df.columns:
                logging.warning("Coluna 'game_id' não encontrada na planilha. Criando-a.")
                sent_games_df['game_id'] = ""

            if 'Tipo de Alerta' not in sent_games_df.columns:
                logging.warning("Coluna 'Tipo de Alerta' não encontrada na planilha. Criando-a.")
                sent_games_df['Tipo de Alerta'] = ""

            logging.info(f"Planilha carregada: {SENT_GAMES_FILE}")

            sent_games_df.to_excel(SENT_GAMES_FILE, index=False)
            format_links_in_excel()
    except Exception as e:
        logging.error(f"Erro ao inicializar a planilha: {e}")
        sent_games_df = pd.DataFrame(columns=[
            'game_id', 'Liga', 'Team A', 'Team B', 'Placar', 'Tempo', 'Odd', 'Linha de gol over', 'Drop', 'Link', 'Tipo de Alerta'
        ])
        sent_games_df.to_excel(SENT_GAMES_FILE, index=False)
        logging.info("Nova planilha criada devido a erro.")

def add_game_to_excel(alert, match_url):
    global sent_games_df
    teams = alert['teams'].split(' - ') if ' - ' in alert['teams'] else [alert['teams'], 'N/A']

    if alert.get('game_id', 'N/A') in sent_games_df['game_id'].values:
        logging.info(f"Jogo com game_id {alert['game_id']} já está na planilha. Ignorando.")
        return

    new_row = {
        'game_id': alert.get('game_id', 'N/A'),
        'Liga': alert['league'],
        'Team A': teams[0],
        'Team B': teams[1],
        'Placar': alert['score'],
        'Tempo': alert['time'],
        'Odd': alert['over'],
        'Linha de gol over': alert['handicap'],
        'Drop': alert['drop'],
        'Link': match_url,
        'Tipo de Alerta': alert.get('alert_type', 'N/A')
    }
    sent_games_df = pd.concat([sent_games_df, pd.DataFrame([new_row])], ignore_index=True)
    sent_games_df.to_excel(SENT_GAMES_FILE, index=False)
    format_links_in_excel()
    logging.info(f"Jogo adicionado à planilha: {new_row}")

def format_links_in_excel():
    wb = load_workbook(SENT_GAMES_FILE)
    ws = wb.active

    link_col = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == 'Link':
            link_col = col
            break

    if link_col:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=link_col)
            if cell.value and not cell.hyperlink:
                cell.hyperlink = Hyperlink(ref=cell.coordinate, target=cell.value)
                cell.font = Font(color="0000FF", underline="single")

    wb.save(SENT_GAMES_FILE)
    logging.info("Links formatados como clicáveis na planilha.")